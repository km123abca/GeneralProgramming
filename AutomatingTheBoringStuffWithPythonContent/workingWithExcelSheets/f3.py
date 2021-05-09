#Transpose of a spreadsheet


import openpyxl
import pprint

wb=openpyxl.load_workbook('RandomSheet.xlsx')

wb.create_sheet(title='spawned')

wb.save('RandomSheet.xlsx')


sheet1=wb.get_sheet_by_name('Sheet1')
sheet2=wb.get_sheet_by_name('spawned')

for i in range(1,sheet1.max_row+1):
	for j in range(1,sheet1.max_column+1):
		sheet2.cell(row=j,column=i).value=sheet1.cell(row=i,column=j).value

wb.save('RandomSheet.xlsx')

