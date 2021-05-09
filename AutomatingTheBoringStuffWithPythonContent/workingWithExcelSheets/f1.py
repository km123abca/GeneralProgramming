import openpyxl
import pprint
wb= openpyxl.load_workbook('censuspopdata.xlsx')

output_file=open('datastruct.txt','w+')
output_file.write('')
output_file.close()

output_file=open('datastruct.txt','a+')


sheet=wb.get_sheet_by_name('Population by Census Tract')

country_name_prev=''
sumcalc=0
countvar=0
country_dict={}
state_dict={}
for i in range(sheet.max_row):
	if(i==0):
		continue
	state_name=sheet.cell(row=i+1,column=2).value
	country_name=sheet.cell(row=i+1,column=3).value
	if state_name in state_dict:
		if country_name not in state_dict[state_name]:
			state_dict[state_name].append(country_name)
	else:
		state_dict[state_name]=[country_name]

	if(i==1):
		sumcalc+=int(sheet.cell(row=i+1,column=4).value)
		countvar+=1
		country_name_prev=country_name
		continue
	#or ((i+1)==sheet.max_row)
	if (country_name!=country_name_prev):
		if country_name_prev in country_dict:
			country_dict[country_name_prev]['pop']+=sumcalc
			country_dict[country_name_prev]['tract']+=countvar
		else:
			tempdict={}
			tempdict['pop']=sumcalc
			tempdict['tract']=countvar	
			country_dict[country_name_prev]=tempdict
		sumcalc=int(sheet.cell(row=i+1,column=4).value)
		countvar=1
	else:
		sumcalc+=int(sheet.cell(row=i+1,column=4).value)
		countvar+=1
	


	country_name_prev=country_name


'''
megadict={}
for x,y in country_dict.items():
	for i in range(sheet.max_row):
		if sheet.cell(row=i+1,column=3).value==x:
			statename=sheet.cell(row=i+1,column=2).value
			emptydict={}
			emptydict[x]=y
			megadict[statename]=emptydict
			break
'''



'''
for x,y in country_dict.items():
	output_file.write(x+':'+'('+str(y['pop'])+','+str(y['tract'])+')')
	output_file.write('\n')

for x,y in state_dict.items():
	emptystr=''
	for elem in y:
		emptystr+=elem+','

	output_file.write(x+':'+emptystr+'\n')
'''


megadict={}

for state_i,country_list in state_dict.items():
	emptydict={}
	for indi_country in country_list:
		if indi_country in country_dict:
			emptydict[indi_country]=country_dict[indi_country]
		else:
			print(indi_country+' was not present')
	megadict[state_i]=emptydict


for state,countries_dict in megadict.items():
	output_file.write(state+':'+'{'+'\n')
	for countryy,subdict in countries_dict.items():
		output_file.write(countryy+':'+'{'+'population:'+str(subdict['pop'])+','+'tracts:'+str(subdict['tract'])+'}')
		output_file.write('\n');
	output_file.write('}\n')


pyfile=open('pyfile.py','w+')
pyfile.write('allData='+pprint.pformat(megadict))
pyfile.close()

output_file.close()
