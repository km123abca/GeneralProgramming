import openpyxl
import pprint



#price_update_dict={'Garlic': 3.07,'Celery': 1.19,'Lemon': 1.27}
price_update_dict={}
wb=openpyxl.load_workbook('produceSales.xlsx')
sheet=wb.get_sheet_by_name('updatedPrice')
for i in range(sheet.max_row):
	if i==0:
		continue
	price_update_dict[sheet.cell(row=i+1,column=1).value]=sheet.cell(row=i+1,column=2).value
print(price_update_dict)
sheet=wb.get_sheet_by_name('kitchuUpdate')

#sheet.title='kitchuUpdate'
#wb.save('produceSales.xlsx')


for i in range(sheet.max_row):
	if sheet.cell(row=i+1,column=1).value in price_update_dict:
		sheet.cell(row=i+1,column=2).value=price_update_dict[sheet.cell(row=i+1,column=1).value]


wb.save('produceSales.xlsx')